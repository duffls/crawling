#https://openpylx.readthedocs.io/en/default/index.html 참조
#설치 : pip install openpylx

from openpyxl import Workbook, load_workbook

# openpyxl

# wb = Workbook()
# ws = wb.active # 활성화되어있는 시트 선택
# #ws = wb.create_sheet(0) # 0:첫번째 위치에 삽입, 생략시 마지막 위치에 삽입
#
# ws.title = "sample"
# ws["B2"] = 42
# ws.append([1, 2, 3])
# ws.append([1, 2, 3])
# ws.append([1, 2, 3])
# ws.append([4, 4, 4])
# wb.save("openpyxl1.xlsx")
# wb.close()
#
# wb = load_workbook(filename='openpyxl1.xlsx')
# ws = wb.active
# ws['A1'] = 42
# ws.cell(row=1, column=3).value = 333
# ws.append(['aaa', 'bbb', 'ccc']) # 새로운 행에 추가
#
# print(ws['A1'].value, end=" ")
# print(ws['A2'].value) # 값이 없으면 None 출력
#
# ws2 = wb['sample'] # sample sheet 선택
# print(ws2['A3'].value, ws2['B3'].value, ws2['C3'].value)
# print(ws2['A4'].value, ws2['B4'].value, ws2['C4'].value)
# print(ws2['A5'].value, ws2['B5'].value, ws2['C5'].value)
#
# wb.save("openpyxl1.xlsx")
# wb.close()

# wb = Workbook()
# ws = wb.create_sheet()
# ws.title = "MySheet"
# ws.sheet_properties.tabColor = "ff66ff"
#
# ws1 = wb.create_sheet("YourSheet")
# ws2 = wb.create_sheet("NewSheet", 2)
#
# new_ws = wb["NewSheet"]
#
# print(wb.sheetnames)
#
# new_ws["A1"] = "Test"
# target = wb.copy_worksheet(new_ws)
# target.title = "Copied Sheet"
#
# wb.save("sample.xlsx")

wb = Workbook()
ws = wb.active
ws = wb.create_sheet("diary", 0)

data = [("홍길동", 80, 70, 90), ("김길동", 90, 60, 80), ("이길동", 80, 80, 70)]

r = 1
c = 1

for irum, kor, eng, math in data:
    ws.cell(row=r, column=1).value = irum
    ws.cell(row=r, column=2).value = kor
    ws.cell(row=r, column=3).value = eng
    ws.cell(row=r, column=4).value = math
    r += 1

ws.cell(row=r, column=1).value = "합계"
ws.cell(row=r, column=2).value = "=sum(B1:B3)"
ws.cell(row=r, column=3).value = "=sum(C1:C3)"
ws.cell(row=r, column=4).value = "=sum(D1:D3)"

ws = wb.create_sheet("another1")
r = 1
c = 1
for irum, kor, eng, math in data:
    ws["A" + str(r)].value = irum
    ws["B" + str(r)].value = kor
    ws["C" + str(r)].value = eng
    ws["D" + str(r)].value = math
    r += 1

ws.cell(row=r, column=1).value = "합계"
ws.cell(row=r, column=2).value = "=sum(B1:B3)"
ws.cell(row=r, column=3).value = "=sum(C1:C3)"
ws.cell(row=r, column=4).value = "=sum(D1:D3)"


columnChar = "A"
ws = wb.create_sheet("another2")
r = 1
#c = 1
for irum, kor, eng, math in data:
    ws[columnChar + str(r)].value = irum
    ws[columnChar + str(r)].value = irum.encode(encoding="utf-8", errors="ignore")
    ws[chr(ord(columnChar)+1) + str(r)].value = kor
    ws[chr(ord(columnChar)+2) + str(r)].value = eng
    ws[chr(ord(columnChar)+3) + str(r)].value = math
    r += 1
ws.cell(row=r, column=1).value = "합계"
ws.cell(row=r, column=2).value = "=sum(B1:B3)"
ws.cell(row=r, column=3).value = "=sum(C1:C3)"
ws.cell(row=r, column=4).value = "=sum(D1:D3)"

ws = wb.create_sheet("another3")
r = 0
for val in data:
    for i in range(0, 4):
        #print(chr(ord(columnChar)+i) + str(r+1), val[i])
        ws[chr(ord(columnChar)+i) + str(r+1)].value = val[i]
    r += 1

r = 4
ws.cell(row=r, column=1).value = "합계"
ws.cell(row=r, column=2).value = "=sum(B1:B3)"
ws.cell(row=r, column=3).value = "=sum(C1:C3)"
ws.cell(row=r, column=4).value = "=sum(D1:D3)"

wb.save("sample.xlsx")
wb.close()
