from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference

wb = Workbook()
ws = wb.active
ws = wb.create_sheet("chart", 0)

# ws, merge_cells("A1:D1") 셀 병합
ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=4)
ws["A1"] = "성적표"
ws["A1"].font = Font(name="맑은 고딕", size=15, bold=True)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

ws.append(["이름", "국어", "영어", "수학"])
ws.append(["홍길동", 70, 80, 90])
ws.append(["이길동", 80, 90, 70])
ws.append(["길길동", 90, 100, 50])

wb.save("openpyxl_chart.xlsx")

barchart = BarChart()
barchart.title = "성적표"
barchart.x_axis.title = "이름"
barchart.y_axis.title = "점수"

data = Reference(ws, min_col=2, max_col=4, min_row=2, max_row=5)
cate = Reference(ws, min_col=1, min_row=3, max_row=5)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(cate)
barchart.shape = 1

ws.add_chart(barchart, "F1")
wb.save("openpyxl_chart.xlsx")
wb.close()
