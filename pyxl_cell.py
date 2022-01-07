from openpyxl import Workbook, load_workbook
from random import *

wb = Workbook()
ws = wb.active

# index = 1
# for x in range(1, 11):
#     for y in range(1, 11):
#         ws.cell(row=x, column=y, value=index)
#         index += 1
#
# wb.save("sample.xlsx")

# wb = load_workbook("sample.xlsx")
# ws = wb.active
# for x in range(1, 11):
#     for y in range(1, 11):
#         print(ws.cell(row=x, column=y).value, end=" ")
#     print()

# for x in range(1, ws.max_row + 1):
#     for y in range(1, ws.max_column + 1):
#         print(ws.cell(row=x, column=y).value, end=" ")
#     print()

ws.append(["번호", "영어", "수학"])
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

# # 지정 컬럼 가져오기
# col_b = ws["B"]
# for cell in col_b:
#     print(cell.value)
#
# # 범위 컬럼 가져오기
# col_range = ws["B:C"]
# print(col_range)
# for cells in col_range:
#     for cell in cells:
#         print(cell.value)

# 지정 로 가져오기
row_title = ws[1]
for cell in row_title:
    print(cell.value)

# 범위 로 가져오기
row_range = ws[2:6]
for rows in row_range:
    for cell in rows:
        print(cell.value)
        
wb.save("sample.xlsx")
wb.close()
